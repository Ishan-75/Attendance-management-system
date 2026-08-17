import csv
import io
import json
from datetime import date, datetime, timezone
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session
from sqlalchemy import and_
from app.models.attendance import Attendance, AttendanceStatus
from app.models.employee import Employee
from app.models.department import Department
from app.schemas.report import AttendanceReportRow, ReportFilter, EmployeeAttendanceSummaryRow

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportService:
    @staticmethod
    def get_attendance_report(
        db: Session,
        filters: ReportFilter
    ) -> List[AttendanceReportRow]:
        """Fetch filtered attendance records for reports ordered department-wise."""
        query = db.query(Attendance).join(Employee).outerjoin(Department, Employee.department_id == Department.id)

        query = query.filter(
            Attendance.attendance_date >= filters.start_date,
            Attendance.attendance_date <= filters.end_date,
            Employee.deleted_at.is_(None)
        )

        if filters.department_id:
            query = query.filter(Employee.department_id == filters.department_id)

        if filters.employee_id:
            query = query.filter(Employee.id == filters.employee_id)

        if filters.status:
            query = query.filter(Attendance.status == filters.status)

        records = query.order_by(
            Department.name.asc().nulls_last(),
            Employee.employee_id.asc(),
            Attendance.attendance_date.desc()
        ).all()

        rows = []
        for r in records:
            emp = r.employee
            dept_name = emp.department.name if emp and emp.department else "General / Unassigned"
            rows.append(
                AttendanceReportRow(
                    employee_code=emp.employee_id if emp else "N/A",
                    employee_name=emp.full_name if emp else "N/A",
                    department_name=dept_name,
                    designation=emp.designation if emp and emp.designation else "Staff",
                    attendance_date=r.attendance_date,
                    status=r.status,
                    check_in=r.check_in_time.strftime("%H:%M") if r.check_in_time else "-",
                    check_out=r.check_out_time.strftime("%H:%M") if r.check_out_time else "-",
                    total_hours=r.total_hours or 0.0,
                    overtime_hours=r.overtime_hours or 0.0,
                    late_minutes=r.late_minutes or 0,
                    early_departure_minutes=r.early_departure_minutes or 0,
                    remarks=r.remarks or ""
                )
            )
        return rows

    @staticmethod
    def export_attendance_csv(
        db: Session,
        filters: ReportFilter
    ) -> io.StringIO:
        """Generate formatted CSV export of attendance report organized by department."""
        rows = ReportService.get_attendance_report(db, filters)

        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

        # Header
        writer.writerow([
            "Department",
            "Employee ID",
            "Employee Name",
            "Designation",
            "Date",
            "Status",
            "Punch In",
            "Punch Out",
            "Total Hours",
            "Overtime (Hrs)",
            "Late (Mins)",
            "Early Departure (Mins)",
            "Remarks"
        ])

        for row in rows:
            writer.writerow([
                row.department_name,
                row.employee_code,
                row.employee_name,
                row.designation,
                row.attendance_date.strftime("%Y-%m-%d"),
                row.status,
                row.check_in,
                row.check_out,
                f"{row.total_hours:.2f}",
                f"{row.overtime_hours:.2f}",
                row.late_minutes,
                row.early_departure_minutes,
                row.remarks
            ])

        output.seek(0)
        return output

    @staticmethod
    def export_attendance_excel(
        db: Session,
        filters: ReportFilter
    ) -> io.BytesIO:
        """
        Generate a premium, Department-Wise grouped Excel (.xlsx) workbook.
        Features:
        - Department Section Grouping with colored headers
        - Individual Employee Punch-In & Punch-Out records
        - Department Subtotals (Records, Worked Hours, Overtime)
        - Grand Organization Total Row
        - Sheet 2: Department KPI Summary Dashboard
        """
        rows = ReportService.get_attendance_report(db, filters)

        wb = openpyxl.Workbook()
        
        # -------------------------------------------------------------
        # SHEET 1: Department-Wise Punch & Attendance Statement
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Department Attendance"
        ws1.views.sheetView[0].showGridLines = True

        # Color & Font Definitions
        title_font = Font(name="Calibri", size=15, bold=True, color="1E3A8A")
        sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        
        dept_header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        dept_header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
        
        col_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        col_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Royal Blue
        
        subtotal_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
        subtotal_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Light Slate
        
        grand_total_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        grand_total_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Midnight Navy
        
        data_font = Font(name="Calibri", size=10, color="1E293B")
        mono_font = Font(name="Consolas", size=9, color="334155")
        
        border_thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )
        
        border_double_bottom = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="double", color="0F172A")
        )

        status_fills = {
            "PRESENT": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            "ABSENT": PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid"),
            "LEAVE": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "HALF_DAY": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
            "WEEK_OFF": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
            "HOLIDAY": PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"),
            "WORK_FROM_HOME": PatternFill(start_color="CFFAFE", end_color="CFFAFE", fill_type="solid"),
        }

        # 1. Title Banner
        ws1["A1"] = "WorkforceHub — Department Attendance & Punch Statement"
        ws1["A1"].font = title_font
        ws1["A2"] = f"Period: {filters.start_date.strftime('%d-%b-%Y')} to {filters.end_date.strftime('%d-%b-%Y')} | Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')} | Total Records: {len(rows)}"
        ws1["A2"].font = sub_font
        ws1.row_dimensions[1].height = 24
        ws1.row_dimensions[2].height = 18

        headers = [
            "Employee ID",
            "Employee Name",
            "Designation",
            "Date",
            "Status",
            "Punch In",
            "Punch Out",
            "Total Hours",
            "Overtime (Hrs)",
            "Late (Mins)",
            "Early Dep (Mins)",
            "Remarks / Notes"
        ]

        # Group rows by Department
        dept_groups: Dict[str, List[AttendanceReportRow]] = {}
        for r in rows:
            dname = r.department_name or "General / Unassigned"
            if dname not in dept_groups:
                dept_groups[dname] = []
            dept_groups[dname].append(r)

        current_row = 4
        grand_total_hours = 0.0
        grand_total_ot = 0.0

        if not dept_groups:
            # Handle empty state
            ws1.cell(row=current_row, column=1, value="No attendance records found for the selected period.").font = sub_font
            current_row += 2
        else:
            for dept_name, dept_rows in dept_groups.items():
                dept_hours_sum = sum(r.total_hours for r in dept_rows)
                dept_ot_sum = sum(r.overtime_hours for r in dept_rows)
                grand_total_hours += dept_hours_sum
                grand_total_ot += dept_ot_sum

                # --- Department Banner Header ---
                ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                dept_banner_cell = ws1.cell(row=current_row, column=1)
                dept_banner_cell.value = f"  🏢 DEPARTMENT: {dept_name.upper()}  ({len(dept_rows)} Records)"
                dept_banner_cell.font = dept_header_font
                dept_banner_cell.fill = dept_header_fill
                dept_banner_cell.alignment = Alignment(horizontal="left", vertical="center")
                ws1.row_dimensions[current_row].height = 26

                for col_idx in range(1, len(headers) + 1):
                    ws1.cell(row=current_row, column=col_idx).border = border_thin
                current_row += 1

                # --- Column Headers for this department ---
                ws1.row_dimensions[current_row].height = 22
                for col_idx, h_text in enumerate(headers, 1):
                    c = ws1.cell(row=current_row, column=col_idx, value=h_text)
                    c.font = col_header_font
                    c.fill = col_header_fill
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = border_thin
                current_row += 1

                # --- Employee Records & Punches ---
                for r in dept_rows:
                    ws1.row_dimensions[current_row].height = 20
                    
                    c1 = ws1.cell(row=current_row, column=1, value=r.employee_code)
                    c1.font = mono_font
                    c1.alignment = Alignment(horizontal="center", vertical="center")

                    c2 = ws1.cell(row=current_row, column=2, value=r.employee_name)
                    c2.font = data_font
                    c2.alignment = Alignment(horizontal="left", vertical="center")

                    c3 = ws1.cell(row=current_row, column=3, value=r.designation)
                    c3.font = data_font
                    c3.alignment = Alignment(horizontal="left", vertical="center")

                    c4 = ws1.cell(row=current_row, column=4, value=r.attendance_date.strftime("%Y-%m-%d"))
                    c4.font = mono_font
                    c4.alignment = Alignment(horizontal="center", vertical="center")

                    c5 = ws1.cell(row=current_row, column=5, value=r.status)
                    c5.font = Font(name="Calibri", size=9, bold=True)
                    c5.alignment = Alignment(horizontal="center", vertical="center")
                    if r.status in status_fills:
                        c5.fill = status_fills[r.status]

                    # Punch In / Check-In
                    c6 = ws1.cell(row=current_row, column=6, value=r.check_in)
                    c6.font = mono_font
                    c6.alignment = Alignment(horizontal="center", vertical="center")

                    # Punch Out / Check-Out
                    c7 = ws1.cell(row=current_row, column=7, value=r.check_out)
                    c7.font = mono_font
                    c7.alignment = Alignment(horizontal="center", vertical="center")

                    # Hours & Overtime
                    c8 = ws1.cell(row=current_row, column=8, value=round(r.total_hours, 2))
                    c8.font = data_font
                    c8.alignment = Alignment(horizontal="right", vertical="center")
                    c8.number_format = "0.00"

                    c9 = ws1.cell(row=current_row, column=9, value=round(r.overtime_hours, 2))
                    c9.font = data_font
                    c9.alignment = Alignment(horizontal="right", vertical="center")
                    c9.number_format = "0.00"

                    c10 = ws1.cell(row=current_row, column=10, value=r.late_minutes if r.late_minutes > 0 else "-")
                    c10.font = data_font
                    c10.alignment = Alignment(horizontal="right", vertical="center")

                    c11 = ws1.cell(row=current_row, column=11, value=r.early_departure_minutes if r.early_departure_minutes > 0 else "-")
                    c11.font = data_font
                    c11.alignment = Alignment(horizontal="right", vertical="center")

                    c12 = ws1.cell(row=current_row, column=12, value=r.remarks)
                    c12.font = sub_font
                    c12.alignment = Alignment(horizontal="left", vertical="center")

                    for col_idx in range(1, len(headers) + 1):
                        ws1.cell(row=current_row, column=col_idx).border = border_thin
                    
                    current_row += 1

                # --- Department Subtotal Row ---
                ws1.row_dimensions[current_row].height = 22
                ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                sub_label_cell = ws1.cell(row=current_row, column=1)
                sub_label_cell.value = f"  Subtotal: {dept_name} ({len(dept_rows)} Records)"
                sub_label_cell.font = subtotal_font
                sub_label_cell.alignment = Alignment(horizontal="left", vertical="center")

                c_sub_h = ws1.cell(row=current_row, column=8, value=round(dept_hours_sum, 2))
                c_sub_h.font = subtotal_font
                c_sub_h.alignment = Alignment(horizontal="right", vertical="center")
                c_sub_h.number_format = "0.00"

                c_sub_ot = ws1.cell(row=current_row, column=9, value=round(dept_ot_sum, 2))
                c_sub_ot.font = subtotal_font
                c_sub_ot.alignment = Alignment(horizontal="right", vertical="center")
                c_sub_ot.number_format = "0.00"

                for col_idx in range(1, len(headers) + 1):
                    cell = ws1.cell(row=current_row, column=col_idx)
                    cell.fill = subtotal_fill
                    cell.border = border_thin

                current_row += 2  # Leave an aesthetic gap between departments

            # --- Grand Total Footer Row ---
            ws1.row_dimensions[current_row].height = 26
            ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            grand_label_cell = ws1.cell(row=current_row, column=1)
            grand_label_cell.value = f"  ★ GRAND TOTAL: ALL DEPARTMENTS ({len(rows)} Total Records)"
            grand_label_cell.font = grand_total_font
            grand_label_cell.alignment = Alignment(horizontal="left", vertical="center")

            c_gt_h = ws1.cell(row=current_row, column=8, value=round(grand_total_hours, 2))
            c_gt_h.font = grand_total_font
            c_gt_h.alignment = Alignment(horizontal="right", vertical="center")
            c_gt_h.number_format = "0.00"

            c_gt_ot = ws1.cell(row=current_row, column=9, value=round(grand_total_ot, 2))
            c_gt_ot.font = grand_total_font
            c_gt_ot.alignment = Alignment(horizontal="right", vertical="center")
            c_gt_ot.number_format = "0.00"

            for col_idx in range(1, len(headers) + 1):
                cell = ws1.cell(row=current_row, column=col_idx)
                cell.fill = grand_total_fill
                cell.border = border_double_bottom

        # Column widths for Sheet 1
        col_widths = {
            "A": 16, # ID
            "B": 26, # Name
            "C": 22, # Designation
            "D": 14, # Date
            "E": 16, # Status
            "F": 14, # Punch In
            "G": 14, # Punch Out
            "H": 15, # Total Hours
            "I": 15, # Overtime
            "J": 13, # Late
            "K": 15, # Early Dep
            "L": 28  # Remarks
        }
        for col_letter, width in col_widths.items():
            ws1.column_dimensions[col_letter].width = width

        # -------------------------------------------------------------
        # SHEET 2: Department KPI Summary Dashboard
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Department KPI Summary")
        ws2.views.sheetView[0].showGridLines = True

        ws2["A1"] = "Department Attendance KPI Summary"
        ws2["A1"].font = title_font
        ws2["A2"] = f"Summary Period: {filters.start_date.strftime('%d-%b-%Y')} to {filters.end_date.strftime('%d-%b-%Y')}"
        ws2["A2"].font = sub_font

        kpi_headers = [
            "Department Name",
            "Total Records",
            "Present",
            "Absent",
            "Leave",
            "Half Day",
            "WFH / Other",
            "Attendance Rate (%)",
            "Total Worked Hours",
            "Overtime Hours"
        ]

        ws2.row_dimensions[4].height = 24
        for col_idx, h_text in enumerate(kpi_headers, 1):
            c = ws2.cell(row=4, column=col_idx, value=h_text)
            c.font = col_header_font
            c.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_thin

        kpi_row = 5
        for dept_name, dept_rows in dept_groups.items():
            ws2.row_dimensions[kpi_row].height = 20
            
            p_cnt = sum(1 for r in dept_rows if r.status == AttendanceStatus.PRESENT)
            a_cnt = sum(1 for r in dept_rows if r.status == AttendanceStatus.ABSENT)
            l_cnt = sum(1 for r in dept_rows if r.status == AttendanceStatus.LEAVE)
            hd_cnt = sum(1 for r in dept_rows if r.status == AttendanceStatus.HALF_DAY)
            wfh_cnt = sum(1 for r in dept_rows if r.status in [AttendanceStatus.WORK_FROM_HOME, AttendanceStatus.HOLIDAY, AttendanceStatus.WEEK_OFF])
            
            total_recs = len(dept_rows)
            att_rate = round(((p_cnt + hd_cnt * 0.5 + sum(1 for r in dept_rows if r.status == AttendanceStatus.WORK_FROM_HOME)) / total_recs) * 100, 1) if total_recs > 0 else 0.0
            
            d_hours = sum(r.total_hours for r in dept_rows)
            d_ot = sum(r.overtime_hours for r in dept_rows)

            ws2.cell(row=kpi_row, column=1, value=dept_name).font = Font(name="Calibri", size=10, bold=True)
            ws2.cell(row=kpi_row, column=2, value=total_recs).alignment = Alignment(horizontal="center")
            ws2.cell(row=kpi_row, column=3, value=p_cnt).alignment = Alignment(horizontal="center")
            ws2.cell(row=kpi_row, column=4, value=a_cnt).alignment = Alignment(horizontal="center")
            ws2.cell(row=kpi_row, column=5, value=l_cnt).alignment = Alignment(horizontal="center")
            ws2.cell(row=kpi_row, column=6, value=hd_cnt).alignment = Alignment(horizontal="center")
            ws2.cell(row=kpi_row, column=7, value=wfh_cnt).alignment = Alignment(horizontal="center")
            
            rate_cell = ws2.cell(row=kpi_row, column=8, value=f"{att_rate:.1f}%")
            rate_cell.alignment = Alignment(horizontal="right")
            rate_cell.font = Font(name="Calibri", size=10, bold=True, color="166534" if att_rate >= 80 else "991B1B")

            h_c = ws2.cell(row=kpi_row, column=9, value=round(d_hours, 2))
            h_c.alignment = Alignment(horizontal="right")
            h_c.number_format = "0.00"

            ot_c = ws2.cell(row=kpi_row, column=10, value=round(d_ot, 2))
            ot_c.alignment = Alignment(horizontal="right")
            ot_c.number_format = "0.00"

            for col_idx in range(1, len(kpi_headers) + 1):
                ws2.cell(row=kpi_row, column=col_idx).border = border_thin

            kpi_row += 1

        kpi_widths = {
            "A": 26, "B": 14, "C": 12, "D": 12, "E": 12,
            "F": 12, "G": 14, "H": 20, "I": 18, "J": 16
        }
        for col_letter, width in kpi_widths.items():
            ws2.column_dimensions[col_letter].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_attendance_json(
        db: Session,
        filters: ReportFilter
    ) -> str:
        """Generate structured JSON attendance export grouped by department."""
        rows = ReportService.get_attendance_report(db, filters)
        
        dept_map: Dict[str, List[Dict[str, Any]]] = {}
        records_list: List[Dict[str, Any]] = []
        for r in rows:
            dname = r.department_name or "General / Unassigned"
            if dname not in dept_map:
                dept_map[dname] = []
            row_dict = {
                "employee_code": r.employee_code,
                "employee_name": r.employee_name,
                "department_name": r.department_name,
                "designation": r.designation,
                "attendance_date": r.attendance_date.strftime("%Y-%m-%d"),
                "status": r.status,
                "punch_in": r.check_in,
                "punch_out": r.check_out,
                "total_hours": r.total_hours,
                "overtime_hours": r.overtime_hours,
                "late_minutes": r.late_minutes,
                "early_departure_minutes": r.early_departure_minutes,
                "remarks": r.remarks
            }
            dept_map[dname].append(row_dict)
            records_list.append(row_dict)

        data = {
            "metadata": {
                "title": "WorkforceHub Attendance Statement",
                "start_date": str(filters.start_date),
                "end_date": str(filters.end_date),
                "total_records": len(rows),
                "department_count": len(dept_map)
            },
            "records": records_list,
            "departments": dept_map
        }
        return json.dumps(data, indent=2)

    @staticmethod
    def export_attendance_html(
        db: Session,
        filters: ReportFilter
    ) -> str:
        """Generate printable styled HTML attendance statement organized department-wise."""
        rows = ReportService.get_attendance_report(db, filters)

        dept_groups: Dict[str, List[AttendanceReportRow]] = {}
        for r in rows:
            dname = r.department_name or "General / Unassigned"
            if dname not in dept_groups:
                dept_groups[dname] = []
            dept_groups[dname].append(r)

        dept_sections_html = ""
        for dept_name, dept_rows in dept_groups.items():
            dept_hours = sum(r.total_hours for r in dept_rows)
            dept_ot = sum(r.overtime_hours for r in dept_rows)

            table_rows_html = ""
            for r in dept_rows:
                badge_class = {
                    "PRESENT": "status-present",
                    "ABSENT": "status-absent",
                    "LEAVE": "status-leave",
                    "HALF_DAY": "status-half",
                    "WORK_FROM_HOME": "status-wfh",
                }.get(r.status, "status-other")

                table_rows_html += f"""
                <tr>
                    <td class="mono">{r.employee_code}</td>
                    <td><strong>{r.employee_name}</strong></td>
                    <td>{r.designation}</td>
                    <td class="mono">{r.attendance_date.strftime('%Y-%m-%d')}</td>
                    <td><span class="badge {badge_class}">{r.status}</span></td>
                    <td class="mono">{r.check_in}</td>
                    <td class="mono">{r.check_out}</td>
                    <td class="num">{r.total_hours:.2f}</td>
                    <td class="num">{r.overtime_hours:.2f}</td>
                    <td>{r.remarks}</td>
                </tr>
                """

            dept_sections_html += f"""
            <div class="dept-section">
                <div class="dept-header">
                    <h3>🏢 Department: {dept_name}</h3>
                    <span>{len(dept_rows)} Records | Worked: {dept_hours:.2f}h | OT: {dept_ot:.2f}h</span>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>Employee ID</th>
                            <th>Employee Name</th>
                            <th>Designation</th>
                            <th>Date</th>
                            <th>Status</th>
                            <th>Punch In</th>
                            <th>Punch Out</th>
                            <th>Total Hrs</th>
                            <th>Overtime</th>
                            <th>Remarks</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
            </div>
            """

        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8" />
    <title>WorkforceHub Attendance Report</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 24px; color: #1e293b; }}
        .header {{ border-bottom: 2px solid #2563eb; padding-bottom: 12px; margin-bottom: 24px; }}
        .header h1 {{ margin: 0; color: #1e3a8a; font-size: 24px; }}
        .header p {{ margin: 4px 0 0; color: #64748b; font-size: 13px; }}
        .dept-section {{ margin-bottom: 28px; page-break-inside: avoid; }}
        .dept-header {{ background: #1e293b; color: white; padding: 8px 14px; border-radius: 6px 6px 0 0; display: flex; justify-content: space-between; align-items: center; }}
        .dept-header h3 {{ margin: 0; font-size: 14px; }}
        .dept-header span {{ font-size: 12px; opacity: 0.9; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 0; }}
        th {{ background: #2563eb; color: white; padding: 8px 10px; text-align: left; font-weight: 600; }}
        td {{ padding: 6px 10px; border-bottom: 1px solid #e2e8f0; }}
        tr:hover {{ background: #f8fafc; }}
        .mono {{ font-family: monospace; }}
        .num {{ text-align: right; }}
        .badge {{ padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: bold; text-transform: uppercase; }}
        .status-present {{ background: #dcfce7; color: #15803d; }}
        .status-absent {{ background: #fee2e2; color: #b91c1c; }}
        .status-leave {{ background: #fef3c7; color: #b45309; }}
        .status-half {{ background: #ffedd5; color: #c2410c; }}
        .status-wfh {{ background: #cffafe; color: #0e7490; }}
        .status-other {{ background: #f1f5f9; color: #475569; }}
        @media print {{
            body {{ margin: 0; }}
            .no-print {{ display: none; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>WorkforceHub Attendance Statement</h1>
        <p>Period: {filters.start_date} to {filters.end_date} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
    </div>
    {dept_sections_html if dept_sections_html else '<p>No attendance records found for the selected period.</p>'}
</body>
</html>
"""
        return html
